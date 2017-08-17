# -*- coding: UTF-8 -*-
import sys
import string
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pprint
import mysql.connector
from datetime import datetime

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.get_sheet_by_name(unicode('上架0815', "utf-8"))

# 字段名称
PRD_GIFT_ID = 'gift_id'
PRD_CODE = 'prd_code'
PRD_NAME = 'prd_name'
PRD_POINT = 'origin_point'
PRD_TYPE = 'type'
PRD_EX_TIMES = 'ex_times'
PRD_MONTH_EXCHANGE = 'month_exchange'
PRD_START_TIME = 'start_time'
PRD_END_TIME = 'end_time'
PRD_EX_TYPE = 'exchange_type'
PRD_STOCK_COUNT = 'total_count'
PRD_MODULE = 'module'
PRD_SEQ_NO = 'seq_no'
PRD_STORE = 'store_scope'
PRD_CARD = 'card_id'
PRD_DEL = 'is_del'
PRD_DESC = 'description'

# xlsx字母列对应的字段名
PRD_COLUMN_NAME_DICT = {}
PRD_COLUMN_NAME_DICT['D'] = PRD_GIFT_ID
PRD_COLUMN_NAME_DICT['E'] = PRD_CODE
PRD_COLUMN_NAME_DICT['F'] = PRD_NAME
PRD_COLUMN_NAME_DICT['G'] = PRD_POINT
PRD_COLUMN_NAME_DICT['I'] = PRD_TYPE
PRD_COLUMN_NAME_DICT['J'] = PRD_EX_TIMES
PRD_COLUMN_NAME_DICT['K'] = PRD_MONTH_EXCHANGE
PRD_COLUMN_NAME_DICT['L'] = PRD_START_TIME
PRD_COLUMN_NAME_DICT['M'] = PRD_END_TIME
PRD_COLUMN_NAME_DICT['N'] = PRD_EX_TYPE
PRD_COLUMN_NAME_DICT['O'] = PRD_STOCK_COUNT
PRD_COLUMN_NAME_DICT['Q'] = PRD_MODULE
PRD_COLUMN_NAME_DICT['R'] = PRD_SEQ_NO
PRD_COLUMN_NAME_DICT['S'] = PRD_STORE
PRD_COLUMN_NAME_DICT['T'] = PRD_CARD
PRD_COLUMN_NAME_DICT['U'] = PRD_DEL
PRD_COLUMN_NAME_DICT['V'] = PRD_DESC

def strippedString(str):
	return string.strip(str)

def formatRowDict(dict):
	# 礼品编码去掉空格
	dict[PRD_GIFT_ID] = strippedString(dict[PRD_GIFT_ID])
	# 产品条形码去掉空格
	dict[PRD_CODE] = strippedString(dict[PRD_CODE])
	# 礼品名称去掉空格
	dict[PRD_NAME] = strippedString(dict[PRD_NAME])
	# 常规兑分值转化为整数
	# dict[PRD_POINT] = int(dict[PRD_POINT])
	# 产品类型截取整数
	dict[PRD_TYPE] = int(dict[PRD_TYPE][0])
	# 限购次数转化为整数
	# dict[PRD_EX_TIMES] = int(dict[PRD_EX_TIMES])
	# 是否每月限制截取整数
	dict[PRD_MONTH_EXCHANGE] = int(dict[PRD_MONTH_EXCHANGE][0])
	# 是否快递截取整数
	dict[PRD_EX_TYPE] = int(dict[PRD_EX_TYPE][0])
	# 所在版块截取整数
	dict[PRD_MODULE] = int(dict[PRD_MODULE][0])
	# 所在位置转化为整数
	# dict[PRD_SEQ_NO] = int(dict[PRD_SEQ_NO])
	# 是否指定门店截取整数
	dict[PRD_STORE] = int(dict[PRD_STORE][0])
	# 是否券码截取整数
	dict[PRD_CARD] = int(dict[PRD_CARD][0])
	# 是否上架截取整数
	dict[PRD_DEL] = int(dict[PRD_DEL][0])
	dict[PRD_DESC] = r'<div align="center"><img src="http://cdn-yyj.4000916916.com/wx/wxExchange/prdDescImage/111500000016_20161122.jpg"></div>'
	return dict

# 将某一行的数据构建成一个字典
def buildRowDict(worksheet, row_index):
	row_dict = {}
	for row in worksheet.iter_rows(min_row=row_index, max_row=row_index):
		for cell in row:
			if cell.column in PRD_COLUMN_NAME_DICT:
				row_dict.setdefault(PRD_COLUMN_NAME_DICT[cell.column], cell.value)
	return formatRowDict(row_dict)

tmp = buildRowDict(sheet, 3)
pprint.pprint(tmp)

cnx = mysql.connector.connect(user='root', password='tcbj',
                              host='192.168.103.107',
                              database='wxexchange')
if cnx.is_connected():
	print '已连接数据库'
else:
	print '数据库连接失败'

cursor = cnx.cursor(buffered=True)

def isGiftExist(cur, gid):
	select_stmt = "SELECT * FROM product WHERE gift_id = %(gift_id)s"
	cur.execute(select_stmt, {'gift_id': gid})
	return (cur.rowcount > 0)

def insertNewGift(conn, cur, gift):
	add_gift = ("INSERT INTO product "
                "(gift_id, prd_code, prd_name, origin_point, type, ex_times, start_time, end_time, exchange_type, seq_no, module, description, store_scope, card_id, is_del, month_exchange) "
                "VALUES (%(gift_id)s, %(prd_code)s, %(prd_name)s, %(origin_point)s, %(type)s, %(ex_times)s, %(start_time)s, %(end_time)s, %(exchange_type)s, %(seq_no)s, %(module)s, %(description)s, %(store_scope)s, %(card_id)s, %(is_del)s, %(month_exchange)s)")
	cur.execute(add_gift, gift)
	conn.commit()

def updateStockId(conn, cur, gift):
	print "待开发"

insertNewGift(cnx, cursor, tmp)

cursor.close()
cnx.close()