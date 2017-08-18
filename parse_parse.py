# -*- coding: UTF-8 -*-
import sys
import string
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pprint
import mysql.connector
from datetime import datetime, date
import os

# ------------------------------------------------------------------------------
# 常量定义
# ------------------------------------------------------------------------------
# 数据库字段名称
ROW_ID = 'id'
ROW_GIFT_ID = 'gift_id'
ROW_PRD_CODE = 'prd_code'
ROW_PRD_NAME = 'prd_name'
ROW_ORIGIN_POINT = 'origin_point'
ROW_TYPE = 'type'
ROW_EX_TIMES = 'ex_times'
ROW_MONTH_EXCHANGE = 'month_exchange'
ROW_START_TIME = 'start_time'
ROW_END_TIME = 'end_time'
ROW_EXCHANGE_TYPE = 'exchange_type'
ROW_TOTAL_COUNT = 'total_count'
ROW_MODULE = 'module'
ROW_SEQ_NO = 'seq_no'
ROW_STORE_SCOPE = 'store_scope'
ROW_CARD_ID = 'card_id'
ROW_IS_DEL = 'is_del'
ROW_DESCRIPTION = 'description'
ROW_STOCK_ID = 'stock_id'

# Excel表字母列对应的数据库字段名
XLS_COLUMN_DICT = {}
XLS_COLUMN_DICT['D'] = ROW_GIFT_ID
XLS_COLUMN_DICT['E'] = ROW_PRD_CODE
XLS_COLUMN_DICT['F'] = ROW_PRD_NAME
XLS_COLUMN_DICT['G'] = ROW_ORIGIN_POINT
XLS_COLUMN_DICT['I'] = ROW_TYPE
XLS_COLUMN_DICT['J'] = ROW_EX_TIMES
XLS_COLUMN_DICT['K'] = ROW_MONTH_EXCHANGE
XLS_COLUMN_DICT['L'] = ROW_START_TIME
XLS_COLUMN_DICT['M'] = ROW_END_TIME
XLS_COLUMN_DICT['N'] = ROW_EXCHANGE_TYPE
XLS_COLUMN_DICT['O'] = ROW_TOTAL_COUNT
XLS_COLUMN_DICT['Q'] = ROW_MODULE
XLS_COLUMN_DICT['R'] = ROW_SEQ_NO
XLS_COLUMN_DICT['S'] = ROW_STORE_SCOPE
XLS_COLUMN_DICT['T'] = ROW_CARD_ID
XLS_COLUMN_DICT['U'] = ROW_IS_DEL
XLS_COLUMN_DICT['V'] = ROW_DESCRIPTION

# ------------------------------------------------------------------------------
# 函数接口：处理Excel表数据
# ------------------------------------------------------------------------------
# 用礼品的条形码构建其描述内容
def buildGiftDesc(code):
	url1 = r'<div align="center"><img src="http://cdn-yyj.4000916916.com/wx/wxExchange/prdDescImage/'
	url2 = "{}_{}".format(code, date.today().strftime("%Y%m%d"))
	url3 = r'.jpg"></div>'
	return (url1+url2+url3)

# 去除字符串左右两侧的空格
def strippedString(str):
	return string.strip(str)

# 格式化礼品记录数据
def formatRowDict(dict):
	# 礼品编码去掉空格
	dict[ROW_GIFT_ID] = strippedString(dict[ROW_GIFT_ID])
	# 产品条形码去掉空格
	dict[ROW_PRD_CODE] = strippedString(dict[ROW_PRD_CODE])
	# 礼品名称去掉空格
	dict[ROW_PRD_NAME] = strippedString(dict[ROW_PRD_NAME])
	# 常规兑分值转化为整数
	# dict[ROW_ORIGIN_POINT] = int(dict[ROW_ORIGIN_POINT])
	# 产品类型截取整数
	dict[ROW_TYPE] = int(dict[ROW_TYPE][0])
	# 限购次数转化为整数
	# dict[ROW_EX_TIMES] = int(dict[ROW_EX_TIMES])
	# 是否每月限制截取整数
	dict[ROW_MONTH_EXCHANGE] = int(dict[ROW_MONTH_EXCHANGE][0])
	# 是否快递截取整数
	dict[ROW_EXCHANGE_TYPE] = int(dict[ROW_EXCHANGE_TYPE][0])
	# 所在版块截取整数
	dict[ROW_MODULE] = int(dict[ROW_MODULE][0])
	# 所在位置转化为整数
	# dict[ROW_SEQ_NO] = int(dict[ROW_SEQ_NO])
	# 是否指定门店截取整数
	dict[ROW_STORE_SCOPE] = int(dict[ROW_STORE_SCOPE][0])
	# 是否券码截取整数
	dict[ROW_CARD_ID] = int(dict[ROW_CARD_ID][0])
	# 是否上架截取整数
	dict[ROW_IS_DEL] = int(dict[ROW_IS_DEL][0])
	dict[ROW_DESCRIPTION] = buildGiftDesc(dict[ROW_PRD_CODE])
	return dict

# 将Excel表某一行的数据构建成一个数据库记录字典
def buildRowDict(worksheet, row_index):
	row_dict = {}
	for row in worksheet.iter_rows(min_row=row_index, max_row=row_index):
		for cell in row:
			if cell.column in XLS_COLUMN_DICT:
				row_dict.setdefault(XLS_COLUMN_DICT[cell.column], cell.value)
	return formatRowDict(row_dict)

# ------------------------------------------------------------------------------
# 函数接口：修改数据库记录
# ------------------------------------------------------------------------------
# 插入一条新的库存记录
def insertStock(cur, total):
	insert_stock = ("INSERT INTO product_stock (total_count) VALUES (%(total_count)s)")
	cur.execute(insert_stock, {'total_count': total})

# 增加库存
def updateStockRow(cursor, stock_id, delta):
	# 查询原来库存总量
	select_stmt = "SELECT total_count FROM product_stock WHERE id = %(stock_id)s"
	cursor.execute(select_stmt, {'stock_id': stock_id})
	if cursor.rowcount > 0:
		# 增加库存后的总量
		total = cursor.fetchone()[ROW_TOTAL_COUNT] + delta
		# 修改库存总量
		update_stmt = "UPDATE product_stock SET total_count = %(total_count)s WHERE id = %(stock_id)s"
		cursor.execute(update_stmt, {'total_count': total, 'stock_id': stock_id})

# 数据库查询在上架状态的礼品记录
def selectGiftRow(cursor, gift_xls):
	select_stmt = "SELECT * FROM product WHERE gift_id = %(gift_id)s AND is_del = 0 AND module = %(module)s"
	cursor.execute(select_stmt, {'gift_id': gift_xls[ROW_GIFT_ID], 'module': gift_xls[ROW_MODULE]})
	gift_row = cursor.fetchone()
	return gift_row

# 插入一条新的礼品记录
def insertGiftRow(cursor, gift_xls):
	# 如果Excel表写的是下架，则不要插入新的记录。
	if gift_xls[ROW_IS_DEL] == 0:
		# 插入礼品
		insert_gift = ("INSERT INTO product "
	                   "(gift_id, prd_code, prd_name, origin_point, type, ex_times, start_time, end_time, exchange_type, seq_no, module, description, store_scope, card_id, is_del, month_exchange) "
	                   "VALUES (%(gift_id)s, %(prd_code)s, %(prd_name)s, %(origin_point)s, %(type)s, %(ex_times)s, %(start_time)s, %(end_time)s, %(exchange_type)s, %(seq_no)s, %(module)s, %(description)s, %(store_scope)s, %(card_id)s, %(is_del)s, %(month_exchange)s)")
		cursor.execute(insert_gift, gift_xls)

		# 修改礼物记录里的库存ID
		update_gift = "UPDATE product SET stock_id = %(stock_id)s WHERE id = %(id)s"
		cursor.execute(update_gift, {'stock_id': cursor.lastrowid, 'id': cursor.lastrowid})

		# 插入一条新的库存记录
		insertStock(cursor, gift_xls[ROW_TOTAL_COUNT])

# 修改礼品记录为下架状态
def unshelveGiftRow(cursor, id):
	update_gift = "UPDATE product SET is_del = %(is_del)s WHERE id = %(id)s"
	cursor.execute(update_gift, {'is_del': 1, 'id': id})

# 修改礼品记录：开始时间、结束时间、增加库存、上下架状态
def updateGiftRow(cursor, gift_row, gift_xls):
	# 修改开始时间、结束时间、上下架状态
	update_gift = "UPDATE product SET start_time = %(start_time)s, end_time = %(end_time)s, is_del = %(is_del)s WHERE id = %(id)s"
	cursor.execute(update_gift, {'start_time': gift_xls[ROW_START_TIME], 'end_time': gift_xls[ROW_END_TIME], 'is_del': gift_xls[ROW_IS_DEL], 'id': gift_row[ROW_ID]})

	# 新增库存总量
	updateStockRow(cursor, gift_row[ROW_STOCK_ID], gift_xls[ROW_TOTAL_COUNT])

# ------------------------------------------------------------------------------
# 业务逻辑：读取Excel表
# ------------------------------------------------------------------------------
workbook = openpyxl.load_workbook(sys.argv[1])
worksheet = workbook.get_sheet_by_name('config')
max_row = worksheet.max_row
# ------------------------------------------------------------------------------
# 业务逻辑：连接数据库
# ------------------------------------------------------------------------------
cnx = mysql.connector.connect(user='root', password='tcbj',
                              host='192.168.103.107',
                              database='wxexchange')
if cnx.is_connected():
	print '++++ 已连接数据库 ++++'

else:
	print '++++ 连接数据库失败，退出程序 ++++'
	os._exit(0)

cursor = cnx.cursor(buffered=True, dictionary=True)

# ------------------------------------------------------------------------------
# 业务逻辑：根据Excel表的配置修改数据库记录
# ------------------------------------------------------------------------------

for row_idx in range(3, max_row):
	gift_xls = buildRowDict(worksheet, row_idx)
	# 查询数据库礼品记录
	gift_row = selectGiftRow(cursor, gift_xls)

	if gift_row: # 数据库礼品记录已存在则根据Excel表修改相关数据
		print "** 礼品已经存在，修改相关数据", gift_xls[ROW_GIFT_ID]
		if gift_row[ROW_MONTH_EXCHANGE] == 1: # 每月兑换有限制
			print "** ** 每月兑换有限制，修改相关数据"
			updateGiftRow(cursor, gift_row, gift_xls)
		else: # 每月兑换无限制
			print "** ** 每月兑换无限制"
			print "**** 目前是上架状态，先修改为下架状态，再插入一条新的记录", gift_xls[ROW_GIFT_ID]
			unshelveGiftRow(cursor, gift_row['id'])
			# 插入一条新记录
			insertGiftRow(cursor, gift_xls)

	else: # 数据库礼品记录不存在则插入一条新记录
		print "-- 礼品不存在，插入一条新记录", gift_xls[ROW_GIFT_ID]
		insertGiftRow(cursor, gift_xls)
	print '++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'

# ------------------------------------------------------------------------------
# 业务逻辑：关闭数据库
# ------------------------------------------------------------------------------
cnx.commit()
cursor.close()
cnx.close()